import os
import sys
import time
import json
import base64
from datetime import datetime
import requests
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
from selenium.webdriver.chrome.service import Service
import google.generativeai as genai
from dotenv import load_dotenv
from pathlib import Path

# Load environment variables and configure Gemini
load_dotenv()
api_key = os.getenv("GEMINI_API_KEY")
if not api_key:
    raise ValueError("No API key found. Please set the GEMINI_API_KEY environment variable.")

genai.configure(api_key=api_key)

generation_config = {
    "temperature": 1,
    "top_p": 0.95,
    "top_k": 40,
    "max_output_tokens": 8192,
}

model = genai.GenerativeModel(
    model_name="gemini-1.5-flash",
    generation_config=generation_config,
)


class Logger:
    def __init__(self, filename):
        self.terminal = sys.stdout
        self.log_file = open(filename, 'a', encoding='utf-8')
        
    def write(self, message):
        timestamp = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        log_message = f"[{timestamp}] {message}"
        self.terminal.write(log_message)
        self.log_file.write(log_message)
        self.log_file.flush()
        
    def flush(self):
        self.terminal.flush()
        self.log_file.flush()

# Redirect stdout to both terminal and log file
log_path = "logs/test.txt"
sys.stdout = Logger(log_path)

# Constants
CONFIG_FILE = "D:\\KarnatakaHC\\test.json"  # Updated to point to the JSON file
WEBSITE_URL = "https://karnatakajudiciary.kar.nic.in/newwebsite/rep_judgment.php"

def load_config():
    with open(CONFIG_FILE, 'r') as f:
        return json.load(f)

def save_config(config):
    with open(CONFIG_FILE, 'w') as f:
        json.dump(config, f, indent=4)

def setup_driver():
    config = load_config()
    
    # Create download directory if it doesn't exist
    download_dir = config["download_directory"]
    if not os.path.exists(download_dir):
        print(f"Creating download directory: {download_dir}")
        os.makedirs(download_dir, exist_ok=True)
    
    chrome_options = webdriver.ChromeOptions()
    
    # Enhanced download preferences
    chrome_options.add_experimental_option('prefs', {
        "download.default_directory": config["download_directory"],
        "download.prompt_for_download": False,
        "download.directory_upgrade": True,
        "safebrowsing.enabled": True,
        "plugins.always_open_pdf_externally": True,
        # Force PDF to download instead of opening in browser
        "download.extensions_to_open": "",
        "browser.download.manager.showWhenStarting": False,
        "browser.download.manager.focusWhenStarting": False,
        "browser.download.manager.useWindow": False,
        "browser.helperApps.neverAsk.saveToDisk": "application/pdf",
        "pdfjs.disabled": True
    })
    
    # Add more compatibility options
    chrome_options.add_argument('--no-sandbox')
    chrome_options.add_argument('--disable-dev-shm-usage')
    chrome_options.add_argument('--disable-gpu')
    chrome_options.add_argument('--disable-software-rasterizer')
    chrome_options.add_argument('--disable-extensions')
    chrome_options.add_experimental_option("detach", True)
    chrome_options.add_argument("--start-maximized")
    chrome_options.add_argument('--ignore-certificate-errors')
    chrome_options.add_argument('--ignore-ssl-errors')
    chrome_options.add_experimental_option('excludeSwitches', ['enable-logging'])
    
    print("Setting up Chrome driver...")
    try:
        # First try with default service
        service = Service()
        driver = webdriver.Chrome(options=chrome_options)
        print("Chrome driver created successfully with default service")
        return driver
    except Exception as e1:
        print(f"Failed with default service: {e1}")
        try:
            # Try with explicit chromedriver path
            service = Service(executable_path=os.path.join(os.getcwd(), "chromedriver.exe"))
            driver = webdriver.Chrome(service=service, options=chrome_options)
            print("Chrome driver created successfully with explicit path")
            return driver
        except Exception as e2:
            print(f"Failed with explicit path: {e2}")
            print("\nTroubleshooting steps:")
            print("1. Make sure Chrome is installed")
            print("2. Download matching chromedriver from: https://googlechromelauncher.github.io/chromedriver/")
            print("3. Place chromedriver.exe in the same folder as this script")
            print("4. Your Chrome version: Check in Chrome menu > Help > About Google Chrome")
            raise

def solve_captcha_with_gemini(driver, wait):
    try:
        time.sleep(2)  # Wait before getting captcha
        
        # Get captcha image
        captcha_img = wait.until(
            EC.presence_of_element_located((By.XPATH, "//img[@id='captcha']"))
        )
        
        # Take screenshot of captcha
        driver.save_screenshot("captcha_temp.png")
        
        # Load the image for Gemini
        with open("captcha_temp.png", "rb") as img_file:
            image_data = {
                "mime_type": "image/png",
                "data": base64.b64encode(img_file.read()).decode('utf-8')
            }

        # Create message parts for Gemini
        message_parts = [
            "This is a captcha image containing exactly 6 digits. Return only these 6 digits, nothing else.",
            image_data
        ]

        # Send to Gemini and get response
        print("Sending captcha to Gemini...")
        response = model.generate_content(message_parts)
        captcha_text = response.text.strip()
        print(f"Gemini identified captcha: {captcha_text}")
        
        if not captcha_text.isdigit() or len(captcha_text) != 6:
            print("Invalid captcha text (not 6 digits)")
            return False
            
        # Enter captcha
        captcha_input = wait.until(
            EC.presence_of_element_located((By.XPATH, "/html/body/div[1]/div[2]/div[1]/form/div[2]/div[2]/div[8]/div[4]/div/input"))
        )
        captcha_input.clear()
        time.sleep(1)
        captcha_input.send_keys(captcha_text)
        time.sleep(2)
        
        # Clean up temporary file
        if os.path.exists("captcha_temp.png"):
            os.remove("captcha_temp.png")
            
        return True
        
    except Exception as e:
        print(f"Error solving captcha: {e}")
        if os.path.exists("captcha_temp.png"):
            os.remove("captcha_temp.png")
        return False

def setup_excel():
    """Setup Excel file with proper headers"""
    config = load_config()
    excel_path = config["excel_path"]
    excel_dir = os.path.dirname(excel_path)
    
    if not os.path.exists(excel_dir):
        os.makedirs(excel_dir, exist_ok=True)
    
    if not os.path.exists(excel_path):
        wb = Workbook()
        ws = wb.active
        headers = ['SNo', 'Case No', 'Year', 'Case Title', 'Decision Date', 
                  'Judge Name', 'PDF Status', 'PDF Filename', 'New Name']
        
        # Add headers and set column widths
        for col, header in enumerate(headers, 1):
            ws.cell(row=1, column=col, value=header)
            ws.column_dimensions[get_column_letter(col)].width = 15
        
        # Make some columns wider
        ws.column_dimensions[get_column_letter(4)].width = 40  # Case Title
        ws.column_dimensions[get_column_letter(8)].width = 30  # PDF Filename
        ws.column_dimensions[get_column_letter(9)].width = 50  # New Name
        
        wb.save(excel_path)
    return excel_path

def update_excel(config, row_num, case_data):
    """Add a row to Excel with case details"""
    try:
        excel_path = config["excel_path"]
        wb = load_workbook(excel_path)
        ws = wb.active
        next_row = ws.max_row + 1
        
        # Write data to row
        ws.cell(row=next_row, column=1, value=row_num)  # SNo
        ws.cell(row=next_row, column=2, value=case_data.get('case_no', ''))
        ws.cell(row=next_row, column=3, value=case_data.get('year', ''))
        ws.cell(row=next_row, column=4, value=case_data.get('case_title', ''))
        ws.cell(row=next_row, column=5, value=case_data.get('decision_date', ''))
        ws.cell(row=next_row, column=6, value=case_data.get('judge_name', ''))
        ws.cell(row=next_row, column=7, value=case_data.get('pdf_status', ''))
        ws.cell(row=next_row, column=8, value=case_data.get('original_filename', ''))
        ws.cell(row=next_row, column=9, value=case_data.get('new_filename', ''))
        
        wb.save(excel_path)
        print(f"Updated Excel for case {row_num}")
    except Exception as e:
        print(f"Error updating Excel for case {row_num}: {e}")

def remove_blocking_elements(driver):
    """Remove elements that might block clicking PDF buttons"""
    try:
        blocker = driver.find_element(By.CLASS_NAME, "swal2-container")
        driver.execute_script("arguments[0].remove();", blocker)
    except:
        pass

def downloaded_pdf_minus1():
    """
    Navigate to the Karnataka HC website, select bench, enter dates, solve CAPTCHA,
    search for results, scroll to the bottom of the page, extract the last serial number,
    and update the end_serial value in the JSON configuration file.
    
    Before starting, it reads status.xlsx to find the first pending task,
    updates the configuration with dates from that task, and later updates
    the status and total_pdfs fields in the Excel file.
    """
    # First check status.xlsx for pending tasks
    status_path = os.path.join(os.path.dirname(CONFIG_FILE), "status.xlsx")
    print(f"Checking status file: {status_path}")
    
    if not os.path.exists(status_path):
        print(f"Error: Status file not found at {status_path}")
        return None
    
    try:
        # Open status.xlsx and find the first pending row
        status_wb = load_workbook(status_path)
        status_ws = status_wb.active
        
        pending_row = None
        for row in range(2, status_ws.max_row + 1):  # Start from row 2 (skip header)
            status_value = status_ws.cell(row=row, column=5).value  # Column E (status)
            if status_value and status_value.lower() == "pending":
                pending_row = row
                break
        
        if pending_row is None:
            print("No pending tasks found in status.xlsx")
            return None
            
        # Get values from the pending row
        start_date = status_ws.cell(row=pending_row, column=2).value  # Column B (start_date)
        end_date = status_ws.cell(row=pending_row, column=3).value  # Column C (end_date)
        month_name = status_ws.cell(row=pending_row, column=4).value  # Column D (month_name)
        
        print(f"Found pending task: {start_date} to {end_date} ({month_name})")
        
        # Format dates for the config file
        if isinstance(start_date, datetime):
            display_from_date = start_date.strftime("%d-%m-%Y")
            from_date = start_date.strftime("%m/%d/%Y")
        else:
            print(f"Warning: start_date not a datetime object: {start_date}")
            from_date = start_date
            display_from_date = start_date
            
        if isinstance(end_date, datetime):
            display_to_date = end_date.strftime("%d-%m-%Y")
            to_date = end_date.strftime("%m/%d/%Y")
        else:
            print(f"Warning: end_date not a datetime object: {end_date}")
            to_date = end_date
            display_to_date = end_date
        
        # Update config with values from status.xlsx
        config = load_config()
        
        # Update date configurations
        config["date_config"]["from_date"] = from_date
        config["date_config"]["to_date"] = to_date
        config["date_config"]["display_from_date"] = display_from_date
        config["date_config"]["display_to_date"] = display_to_date
        config["date_config"]["display_month"] = month_name.lower() if month_name else ""
        
        # Update paths
        year_str = ""
        if isinstance(start_date, datetime):
            year_str = str(start_date.year)
        
        if month_name and year_str:
            # Set download directory and Excel path based on month and year
            config["download_directory"] = f"D:\\DATA\\karnataka\\{year_str}\\{month_name.lower()}"
            config["excel_path"] = f"D:\\DATA\\records\\karnataka_{month_name.lower()}.xlsx"
            
            # Make sure the download directory exists
            os.makedirs(config["download_directory"], exist_ok=True)
            
            print(f"Updated config with download dir: {config['download_directory']}")
            print(f"Updated config with Excel path: {config['excel_path']}")
        
        # Save config before web automation
        save_config(config)
    except Exception as e:
        print(f"Error processing status file: {e}")
        return None
    
    # Now continue with the web automation to get the end_serial
    config = load_config()
    print("Starting automated serial number retrieval...")
    
    try:
        driver = setup_driver()
        wait = WebDriverWait(driver, 20)
        print("Driver setup complete, proceeding to website...")
        
        # Open website
        driver.get(WEBSITE_URL)
        print("Website loaded")
        time.sleep(5)  # Initial wait after page load

        # Select bench with proper delays
        print("Selecting bench...")
        bench_select = wait.until(
            EC.element_to_be_clickable((By.XPATH, "/html/body/div[1]/div[2]/div[1]/div[2]/div[3]/select"))
        )
        time.sleep(2)
        bench_select.click()
        time.sleep(3)
        
        print(f"Selecting {config['bench']}...")
        principal_bench = wait.until(
            EC.element_to_be_clickable((By.XPATH, "/html/body/div[1]/div[2]/div[1]/div[2]/div[3]/select/option[2]"))
        )
        time.sleep(2)
        principal_bench.click()
        time.sleep(3)

        # Enter dates with delays
        print("Entering from date...")
        from_date = wait.until(
            EC.presence_of_element_located((By.XPATH, "/html/body/div[1]/div[2]/div[1]/form/div[2]/div[2]/div[6]/div[2]/div/input"))
        )
        time.sleep(2)
        from_date.clear()
        time.sleep(1)
        from_date.send_keys(config["date_config"]["from_date"])
        time.sleep(2)

        print("Entering to date...")
        to_date = wait.until(
            EC.presence_of_element_located((By.XPATH, "/html/body/div[1]/div[2]/div[1]/form/div[2]/div[2]/div[6]/div[4]/div/input"))
        )
        time.sleep(2)
        to_date.clear()
        time.sleep(1)
        to_date.send_keys(config["date_config"]["to_date"])
        time.sleep(2)

        # Solve captcha
        print("Attempting to solve captcha...")
        if not solve_captcha_with_gemini(driver, wait):
            print("Failed to solve captcha automatically")
            input("Please solve the captcha manually and press Enter to continue...")
            # After manual intervention, give time for page to reload if needed
            time.sleep(3)

        # Click search with delay
        print("Clicking search button...")
        search_button = wait.until(
            EC.element_to_be_clickable((By.XPATH, "/html/body/div[1]/div[2]/div[1]/form/div[2]/div[2]/div[9]/div[2]/button[1]"))
        )
        time.sleep(2)
        search_button.click()
        
        # Add longer wait for search results to load completely
        print("Waiting for search results to load (60 seconds)...")
        time.sleep(60)  # Wait for 60 seconds after search button click
        
        # Scroll to the bottom of the page to ensure all rows are loaded
        print("Scrolling to the bottom of the page...")
        driver.execute_script("window.scrollTo(0, document.body.scrollHeight);")
        time.sleep(5)  # Wait for any dynamic loading to complete
        
        # Find all table rows
        print("Locating table rows...")
        rows = driver.find_elements(By.XPATH, "//table[@id='search-results']/tbody/tr")
        
        if not rows:
            # Try an alternative xpath if the first one doesn't work
            rows = driver.find_elements(By.XPATH, "//div[contains(@class, 'dataTables_wrapper')]/table/tbody/tr")
        
        if not rows:
            # Try another alternative if previous attempts failed
            rows = driver.find_elements(By.XPATH, "//div[1]/div[2]/div[1]/div[4]/div/div/div/div[2]/div/div[2]/table/tbody/tr")
        
        if rows:
            # Get the last row
            last_row = rows[-1]
            
            # Get the serial number from the first cell of the last row
            serial_cell = last_row.find_element(By.XPATH, "./td[1]")
            serial_number_text = serial_cell.text.strip()
            
            try:
                serial_number = int(serial_number_text)
                print(f"Found last serial number: {serial_number}")
                
                # Update the configuration file
                config["pdf_range"]["end_serial"] = serial_number
                config["last_updated"] = datetime.now().strftime("%Y-%m-%d")
                
                # Save the updated configuration
                save_config(config)
                
                # Update status.xlsx with the total_pdfs and change status to "processing"
                status_wb = load_workbook(status_path)
                status_ws = status_wb.active
                
                # Update total_pdfs (column F) and status (column E)
                status_ws.cell(row=pending_row, column=6, value=serial_number)  # Column F (total_pdfs)
                status_ws.cell(row=pending_row, column=5, value="processing")   # Column E (status)
                
                status_wb.save(status_path)
                print(f"Updated status.xlsx: total_pdfs={serial_number}, status=processing")
                
                print(f"Configuration updated: end_serial = {serial_number}")
                return serial_number
                
            except ValueError:
                print(f"Error: Could not convert '{serial_number_text}' to integer")
        else:
            print("No table rows found. Check if the search results loaded correctly.")
            
    except Exception as e:
        print(f"Error in downloaded_pdf_minus1: {e}")
    finally:
        if 'driver' in locals():
            print("Closing browser...")
            driver.quit()
    
    return None

def main():
    config = load_config()
    print("Starting automation...")
    
    # Check if downloaded_pdf_number is -1, which indicates we should update the end_serial first
    if config["downloaded_pdf_number"] == -1:
        print("Downloaded PDF number is -1, running serial number update first...")
        new_end_serial = downloaded_pdf_minus1()
        if new_end_serial:
            print(f"Updated end_serial to: {new_end_serial}")
            # Load config again in case it was updated
            config = load_config()
            # Set downloaded_pdf_number to 0 so we can start downloading from the beginning
            config["downloaded_pdf_number"] = 0
            save_config(config)
            print("Reset downloaded_pdf_number to 0 to start fresh downloads")
        else:
            print("Failed to update end_serial. Exiting.")
            return
    
    # Setup Excel file first
    setup_excel()  # Add this line to create Excel file at start
    
    # Create necessary directories
    download_dir = config["download_directory"]
    excel_dir = os.path.dirname(config["excel_path"])
    
    # Create directories if they don't exist
    for dir_path in [download_dir, excel_dir]:
        if not os.path.exists(dir_path):
            print(f"Creating directory: {dir_path}")
            os.makedirs(dir_path, exist_ok=True)
    
    # Check if we've already downloaded all PDFs for the current batch
    if config["downloaded_pdf_number"] >= config["pdf_range"]["end_serial"]:
        print(f"All PDFs already downloaded (total: {config['pdf_range']['end_serial']})")
        
        # Mark as completed in status.xlsx and reset downloaded_pdf_number
        mark_task_as_completed(config)
        
        # Reset downloaded_pdf_number to -1 to prepare for the next batch
        config["downloaded_pdf_number"] = -1
        save_config(config)
        print("Reset downloaded_pdf_number to -1 for next batch")
        return
    
    try:
        driver = setup_driver()
        wait = WebDriverWait(driver, 20)
        print("Driver setup complete, proceeding to website...")
        
        # Open website
        driver.get(WEBSITE_URL)
        print("Website loaded")
        time.sleep(5)  # Initial wait after page load

        # Select bench with proper delays
        print("Selecting bench...")
        bench_select = wait.until(
            EC.element_to_be_clickable((By.XPATH, "/html/body/div[1]/div[2]/div[1]/div[2]/div[3]/select"))
        )
        time.sleep(2)
        bench_select.click()
        time.sleep(3)
        
        print("Selecting Principal Bench...")
        principal_bench = wait.until(
            EC.element_to_be_clickable((By.XPATH, "/html/body/div[1]/div[2]/div[1]/div[2]/div[3]/select/option[2]"))
        )
        time.sleep(2)
        principal_bench.click()
        time.sleep(3)

        # Enter dates with delays
        print("Entering from date...")
        from_date = wait.until(
            EC.presence_of_element_located((By.XPATH, "/html/body/div[1]/div[2]/div[1]/form/div[2]/div[2]/div[6]/div[2]/div/input"))
        )
        time.sleep(2)
        from_date.clear()
        time.sleep(1)
        from_date.send_keys(config["date_config"]["from_date"])
        time.sleep(2)

        print("Entering to date...")
        to_date = wait.until(
            EC.presence_of_element_located((By.XPATH, "/html/body/div[1]/div[2]/div[1]/form/div[2]/div[2]/div[6]/div[4]/div/input"))
        )
        time.sleep(2)
        to_date.clear()
        time.sleep(1)
        to_date.send_keys(config["date_config"]["to_date"])
        time.sleep(2)

        # Solve captcha
        print("Attempting to solve captcha...")
        if not solve_captcha_with_gemini(driver, wait):
            print("Failed to solve captcha")
            input("Press Enter to try again...")
            return

        # Click search with delay
        print("Clicking search button...")
        search_button = wait.until(
            EC.element_to_be_clickable((By.XPATH, "/html/body/div[1]/div[2]/div[1]/form/div[2]/div[2]/div[9]/div[2]/button[1]"))
        )
        time.sleep(2)
        search_button.click()
        
        # Add longer wait for search results to load completely
        print("Waiting for search results to load (60 seconds)...")
        time.sleep(60)  # Wait for 60 seconds after search button click
        
        print("Starting PDF downloads...")
        start_from = config["downloaded_pdf_number"] + 1
        end_at = config["pdf_range"]["end_serial"]

        for i in range(start_from, end_at + 1):
            try:
                print(f"Processing PDF {i}...")
                time.sleep(2)
                
                # Collect case details first
                case_data = {}
                try:
                    case_data['case_no'] = wait.until(
                        EC.presence_of_element_located((By.XPATH, f"/html/body/div[1]/div[2]/div[1]/div[4]/div/div/div/div[2]/div/div[2]/table/tbody/tr[{i}]/td[3]/button/u"))
                    ).text.strip()
                    
                    case_data['year'] = wait.until(
                        EC.presence_of_element_located((By.XPATH, f"/html/body/div[1]/div[2]/div[1]/div[4]/div/div/div/div[2]/div/div[2]/table/tbody/tr[{i}]/td[4]/button/u"))
                    ).text.strip()
                    
                    case_data['decision_date'] = wait.until(
                        EC.presence_of_element_located((By.XPATH, f"/html/body/div[1]/div[2]/div[1]/div[4]/div/div/div/div[2]/div/div[2]/table/tbody/tr[{i}]/td[9]"))
                    ).text.strip()
                    
                    case_data['judge_name'] = wait.until(
                        EC.presence_of_element_located((By.XPATH, f"/html/body/div[1]/div[2]/div[1]/div[4]/div/div/div/div[2]/div/div[2]/table/tbody/tr[{i}]/td[8]"))
                    ).text.strip()
                    
                    parties = wait.until(
                        EC.presence_of_element_located((By.XPATH, f"/html/body/div[1]/div[2]/div[1]/div[4]/div/div/div/div[2]/div/div[2]/table/tbody/tr[{i}]/td[6]"))
                    ).text.strip()
                    case_data['case_title'] = parties
                    
                except Exception as e:
                    print(f"Error collecting case details for row {i}: {e}")
                    continue
                
                # Get list of files before download
                download_dir = Path(config["download_directory"])
                files_before = set(download_dir.glob('*.pdf'))
                
                # Before clicking PDF button, remove any blocking elements
                remove_blocking_elements(driver)
                
                # Click PDF button
                pdf_xpath = f"/html/body/div[1]/div[2]/div[1]/div[4]/div/div/div/div[2]/div/div[2]/table/tbody/tr[{i}]/td[15]/button"
                pdf_button = wait.until(EC.element_to_be_clickable((By.XPATH, pdf_xpath)))
                
                # Scroll the button into view and ensure it's clickable
                driver.execute_script("arguments[0].scrollIntoView({block: 'center'});", pdf_button)
                time.sleep(1)
                
                # Try JavaScript click if regular click fails
                try:
                    pdf_button.click()
                except:
                    driver.execute_script("arguments[0].click();", pdf_button)
                
                time.sleep(2)
                
                # Check for record not found popup
                try:
                    ok_button = WebDriverWait(driver, 3).until(
                        EC.element_to_be_clickable((By.XPATH, "/html/body/div[9]/div/div[6]/button[1]"))
                    )
                    ok_button.click()
                    time.sleep(1)
                    print(f"Record not found for PDF {i}, clicked OK")
                    
                    # Update Excel with NOT AVAILABLE status
                    case_data['pdf_status'] = 'NOT AVAILABLE'
                    case_data['original_filename'] = ''
                    case_data['new_filename'] = ''
                    update_excel(config, i, case_data)
                    continue
                    
                except:
                    # No popup found, proceed with download and renaming
                    try:
                        # Wait for download to complete (increased max time to 45 seconds)
                        download_complete = False
                        max_wait = 45  # Increased from 30 to 45 seconds
                        print(f"Waiting for PDF {i} to download...")
                        
                        while max_wait > 0 and not download_complete:
                            time.sleep(2)  # Check every 2 seconds instead of 1
                            files_after = set(download_dir.glob('*.pdf'))
                            new_files = files_after - files_before
                            if new_files:
                                print(f"PDF {i} download detected")
                                download_complete = True
                                # Add extra wait to ensure file is completely written
                                time.sleep(3)  # Wait 3 more seconds after detection
                                break
                            max_wait -= 2
                            if max_wait % 10 == 0:  # Log progress every 10 seconds
                                print(f"Still waiting for PDF {i}... {max_wait} seconds remaining")
                        
                        if download_complete:
                            # Get the new file
                            new_file = list(new_files)[0]
                            original_filename = new_file.name
                            
                            # Construct new filename
                            parties_split = parties.split(" VS ", 1)
                            first_party = parties_split[0].strip() if len(parties_split) > 0 else "PARTY1"
                            second_party = parties_split[1].strip() if len(parties_split) > 1 else "PARTY2"
                            
                            new_filename = f"KAHC_{case_data['case_no']}_{case_data['year']}_{first_party}_VS_{second_party}.pdf"
                            new_filename = "".join(c for c in new_filename if c.isalnum() or c in "._- ")
                            
                            # Ensure the source file exists and is complete
                            old_path = download_dir / original_filename
                            new_path = download_dir / new_filename
                            
                            # Wait for file to be ready for renaming
                            time.sleep(2)
                            
                            try:
                                old_path.rename(new_path)
                                print(f"Successfully renamed {original_filename} to {new_filename}")
                                
                                # Update Excel with successful download
                                case_data['pdf_status'] = 'DOWNLOADED'
                                case_data['original_filename'] = original_filename
                                case_data['new_filename'] = new_filename
                                update_excel(config, i, case_data)
                                
                                # Update config
                                config["downloaded_pdf_number"] = i
                                save_config(config)
                            except Exception as rename_error:
                                print(f"Error renaming file: {rename_error}")
                                raise
                        else:
                            print(f"Timeout waiting for PDF download for row {i}")
                            raise Exception("Download timeout")
                            
                    except Exception as e:
                        print(f"Error processing download/rename for PDF {i}: {e}")
                        continue

                # Check if we've reached the end serial
                if i >= config["pdf_range"]["end_serial"]:
                    print(f"\nReached end serial number {config['pdf_range']['end_serial']}")
                    print("Saving final Excel updates...")
                    
                    # Get Excel workbook and save
                    try:
                        wb = load_workbook(config["excel_path"])
                        wb.save(config["excel_path"])
                        print(f"Excel file saved successfully: {config['excel_path']}")
                    except Exception as excel_error:
                        print(f"Error saving final Excel update: {excel_error}")
                    
                    # Mark as completed in status.xlsx and reset downloaded_pdf_number
                    mark_task_as_completed(config)
                    
                    # Reset downloaded_pdf_number to -1 to prepare for the next batch
                    config["downloaded_pdf_number"] = -1
                    save_config(config)
                    print("Reset downloaded_pdf_number to -1 for next batch")
                    
                    print("\nScript completed successfully!")
                    return  # Exit the function after reaching end serial

            except Exception as e:
                print(f"Error downloading PDF {i}: {e}")
                continue

    except Exception as e:
        print(f"Error in main process: {e}")
        input("Press Enter to close the browser...")
    finally:
        if 'driver' in locals():
            driver.quit()

def mark_task_as_completed(config):
    """
    Mark the current task as completed in status.xlsx.
    Finds the row where start_date matches from_date in the config.
    """
    status_path = os.path.join(os.path.dirname(CONFIG_FILE), "status.xlsx")
    print(f"Updating status file: {status_path}")
    
    if not os.path.exists(status_path):
        print(f"Warning: Status file not found at {status_path}")
        return
    
    try:
        # Open status.xlsx
        status_wb = load_workbook(status_path)
        status_ws = status_wb.active
        
        from_date = config["date_config"]["from_date"]
        from_date_formatted = from_date
        
        # Try to convert from_date to a datetime for comparison
        try:
            from_date_dt = datetime.strptime(from_date, "%m/%d/%Y")
        except ValueError:
            try:
                from_date_dt = datetime.strptime(from_date, "%d/%m/%Y")
            except ValueError:
                print(f"Warning: Could not parse from_date: {from_date}")
                from_date_dt = None
        
        matching_row = None
        
        # Find the matching row in status.xlsx
        for row in range(2, status_ws.max_row + 1):  # Start from row 2 (skip header)
            start_date = status_ws.cell(row=row, column=2).value  # Column B (start_date)
            
            # Compare as strings if direct comparison fails
            if start_date and isinstance(start_date, datetime) and from_date_dt:
                if start_date.date() == from_date_dt.date():
                    matching_row = row
                    break
            elif start_date:
                # Try string comparison as fallback
                start_date_str = str(start_date)
                if start_date_str in from_date or from_date in start_date_str:
                    matching_row = row
                    break
                
                # Check status value
                status_value = status_ws.cell(row=row, column=5).value  # Column E (status)
                if status_value and status_value.lower() == "processing":
                    matching_row = row
                    break
        
        if matching_row:
            # Update status to "completed"
            status_ws.cell(row=matching_row, column=5, value="completed")  # Column E (status)
            status_wb.save(status_path)
            print(f"Updated status.xlsx: marked row {matching_row} as completed")
        else:
            print("No matching task found in status.xlsx")
            
    except Exception as e:
        print(f"Error updating status file: {e}")

if __name__ == "__main__":
    print("Script starting...")
    
    # Check if a specific argument was passed
    if len(sys.argv) > 1 and sys.argv[1] == "update-serial":
        print("Running in update-serial mode...")
        downloaded_pdf_minus1()
    else:
        # Run the main process for downloading PDFs
        main()
    
    print("Script finished.")
    input("Press Enter to exit...")

