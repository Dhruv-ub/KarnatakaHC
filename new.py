import os
import sys
import time
import json
import base64
from datetime import datetime
import requests
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
from selenium.webdriver.chrome.service import Service
import google.generativeai as genai
from dotenv import load_dotenv
from pathlib import Path

# Load environment variables and configure Gemini
load_dotenv()
api_key = os.getenv("GEMINI_API_KEY")
if not api_key:
    raise ValueError("No API key found. Please set the GEMINI_API_KEY environment variable.")

genai.configure(api_key=api_key)

generation_config = {
    "temperature": 1,
    "top_p": 0.95,
    "top_k": 40,
    "max_output_tokens": 8192,
}

model = genai.GenerativeModel(
    model_name="gemini-1.5-flash",
    generation_config=generation_config,
)

# Constants
CONFIG_FILE = "new.json"
WEBSITE_URL = "https://karnatakajudiciary.kar.nic.in/newwebsite/rep_judgment.php"

def load_config():
    with open(CONFIG_FILE, 'r') as f:
        return json.load(f)

def save_config(config):
    with open(CONFIG_FILE, 'w') as f:
        json.dump(config, f, indent=4)

def setup_driver():
    config = load_config()
    
    # Create download directory if it doesn't exist
    download_dir = config["download_directory"]
    if not os.path.exists(download_dir):
        print(f"Creating download directory: {download_dir}")
        os.makedirs(download_dir, exist_ok=True)
    
    chrome_options = webdriver.ChromeOptions()
    
    # Enhanced download preferences
    chrome_options.add_experimental_option('prefs', {
        "download.default_directory": config["download_directory"],
        "download.prompt_for_download": False,
        "download.directory_upgrade": True,
        "safebrowsing.enabled": True,
        "plugins.always_open_pdf_externally": True,
        # Force PDF to download instead of opening in browser
        "download.extensions_to_open": "",
        "browser.download.manager.showWhenStarting": False,
        "browser.download.manager.focusWhenStarting": False,
        "browser.download.manager.useWindow": False,
        "browser.helperApps.neverAsk.saveToDisk": "application/pdf",
        "pdfjs.disabled": True
    })
    
    # Add more compatibility options
    chrome_options.add_argument('--no-sandbox')
    chrome_options.add_argument('--disable-dev-shm-usage')
    chrome_options.add_argument('--disable-gpu')
    chrome_options.add_argument('--disable-software-rasterizer')
    chrome_options.add_argument('--disable-extensions')
    chrome_options.add_experimental_option("detach", True)
    chrome_options.add_argument("--start-maximized")
    chrome_options.add_argument('--ignore-certificate-errors')
    chrome_options.add_argument('--ignore-ssl-errors')
    chrome_options.add_experimental_option('excludeSwitches', ['enable-logging'])
    
    print("Setting up Chrome driver...")
    try:
        # First try with default service
        service = Service()
        driver = webdriver.Chrome(options=chrome_options)
        print("Chrome driver created successfully with default service")
        return driver
    except Exception as e1:
        print(f"Failed with default service: {e1}")
        try:
            # Try with explicit chromedriver path
            service = Service(executable_path=os.path.join(os.getcwd(), "chromedriver.exe"))
            driver = webdriver.Chrome(service=service, options=chrome_options)
            print("Chrome driver created successfully with explicit path")
            return driver
        except Exception as e2:
            print(f"Failed with explicit path: {e2}")
            print("\nTroubleshooting steps:")
            print("1. Make sure Chrome is installed")
            print("2. Download matching chromedriver from: https://googlechromelauncher.github.io/chromedriver/")
            print("3. Place chromedriver.exe in the same folder as this script")
            print("4. Your Chrome version: Check in Chrome menu > Help > About Google Chrome")
            raise

def solve_captcha_with_gemini(driver, wait):
    try:
        time.sleep(2)  # Wait before getting captcha
        
        # Get captcha image
        captcha_img = wait.until(
            EC.presence_of_element_located((By.XPATH, "//img[@id='captcha']"))
        )
        
        # Take screenshot of captcha
        driver.save_screenshot("captcha_temp.png")
        
        # Load the image for Gemini
        with open("captcha_temp.png", "rb") as img_file:
            image_data = {
                "mime_type": "image/png",
                "data": base64.b64encode(img_file.read()).decode('utf-8')
            }

        # Create message parts for Gemini
        message_parts = [
            "This is a captcha image containing exactly 6 digits. Return only these 6 digits, nothing else.",
            image_data
        ]

        # Send to Gemini and get response
        print("Sending captcha to Gemini...")
        response = model.generate_content(message_parts)
        captcha_text = response.text.strip()
        print(f"Gemini identified captcha: {captcha_text}")
        
        if not captcha_text.isdigit() or len(captcha_text) != 6:
            print("Invalid captcha text (not 6 digits)")
            return False
            
        # Enter captcha
        captcha_input = wait.until(
            EC.presence_of_element_located((By.XPATH, "/html/body/div[1]/div[2]/div[1]/form/div[2]/div[2]/div[8]/div[4]/div/input"))
        )
        captcha_input.clear()
        time.sleep(1)
        captcha_input.send_keys(captcha_text)
        time.sleep(2)
        
        # Clean up temporary file
        if os.path.exists("captcha_temp.png"):
            os.remove("captcha_temp.png")
            
        return True
        
    except Exception as e:
        print(f"Error solving captcha: {e}")
        if os.path.exists("captcha_temp.png"):
            os.remove("captcha_temp.png")
        return False

def setup_excel():
    """Setup Excel file with proper headers"""
    config = load_config()
    excel_path = config["excel_path"]
    excel_dir = os.path.dirname(excel_path)
    
    if not os.path.exists(excel_dir):
        os.makedirs(excel_dir, exist_ok=True)
    
    if not os.path.exists(excel_path):
        wb = Workbook()
        ws = wb.active
        headers = ['SNo', 'Case No', 'Year', 'Case Title', 'Decision Date', 
                  'Judge Name', 'PDF Status', 'PDF Filename', 'New Name']
        
        # Add headers and set column widths
        for col, header in enumerate(headers, 1):
            ws.cell(row=1, column=col, value=header)
            ws.column_dimensions[get_column_letter(col)].width = 15
        
        # Make some columns wider
        ws.column_dimensions[get_column_letter(4)].width = 40  # Case Title
        ws.column_dimensions[get_column_letter(8)].width = 30  # PDF Filename
        ws.column_dimensions[get_column_letter(9)].width = 50  # New Name
        
        wb.save(excel_path)
    return excel_path

def update_excel(config, row_num, case_data):
    """Add a row to Excel with case details"""
    try:
        excel_path = config["excel_path"]
        wb = load_workbook(excel_path)
        ws = wb.active
        next_row = ws.max_row + 1
        
        # Write data to row
        ws.cell(row=next_row, column=1, value=row_num)  # SNo
        ws.cell(row=next_row, column=2, value=case_data.get('case_no', ''))
        ws.cell(row=next_row, column=3, value=case_data.get('year', ''))
        ws.cell(row=next_row, column=4, value=case_data.get('case_title', ''))
        ws.cell(row=next_row, column=5, value=case_data.get('decision_date', ''))
        ws.cell(row=next_row, column=6, value=case_data.get('judge_name', ''))
        ws.cell(row=next_row, column=7, value=case_data.get('pdf_status', ''))
        ws.cell(row=next_row, column=8, value=case_data.get('original_filename', ''))
        ws.cell(row=next_row, column=9, value=case_data.get('new_filename', ''))
        
        wb.save(excel_path)
        print(f"Updated Excel for case {row_num}")
    except Exception as e:
        print(f"Error updating Excel for case {row_num}: {e}")

def remove_blocking_elements(driver):
    """Remove elements that might block clicking PDF buttons"""
    try:
        blocker = driver.find_element(By.CLASS_NAME, "swal2-container")
        driver.execute_script("arguments[0].remove();", blocker)
    except:
        pass

def main():
    config = load_config()
    print("Starting automation...")
    
    # Setup Excel file first
    setup_excel()  # Add this line to create Excel file at start
    
    # Create necessary directories
    download_dir = config["download_directory"]
    excel_dir = os.path.dirname(config["excel_path"])
    
    # Create directories if they don't exist
    for dir_path in [download_dir, excel_dir]:
        if not os.path.exists(dir_path):
            print(f"Creating directory: {dir_path}")
            os.makedirs(dir_path, exist_ok=True)
    
    try:
        driver = setup_driver()
        wait = WebDriverWait(driver, 20)
        print("Driver setup complete, proceeding to website...")
        
        # Open website
        driver.get(WEBSITE_URL)
        print("Website loaded")
        time.sleep(5)  # Initial wait after page load

        # Select bench with proper delays
        print("Selecting bench...")
        bench_select = wait.until(
            EC.element_to_be_clickable((By.XPATH, "/html/body/div[1]/div[2]/div[1]/div[2]/div[3]/select"))
        )
        time.sleep(2)
        bench_select.click()
        time.sleep(3)
        
        print("Selecting Principal Bench...")
        principal_bench = wait.until(
            EC.element_to_be_clickable((By.XPATH, "/html/body/div[1]/div[2]/div[1]/div[2]/div[3]/select/option[2]"))
        )
        time.sleep(2)
        principal_bench.click()
        time.sleep(3)

        # Enter dates with delays
        print("Entering from date...")
        from_date = wait.until(
            EC.presence_of_element_located((By.XPATH, "/html/body/div[1]/div[2]/div[1]/form/div[2]/div[2]/div[6]/div[2]/div/input"))
        )
        time.sleep(2)
        from_date.clear()
        time.sleep(1)
        from_date.send_keys(config["date_config"]["from_date"])
        time.sleep(2)

        print("Entering to date...")
        to_date = wait.until(
            EC.presence_of_element_located((By.XPATH, "/html/body/div[1]/div[2]/div[1]/form/div[2]/div[2]/div[6]/div[4]/div/input"))
        )
        time.sleep(2)
        to_date.clear()
        time.sleep(1)
        to_date.send_keys(config["date_config"]["to_date"])
        time.sleep(2)

        # Solve captcha
        print("Attempting to solve captcha...")
        if not solve_captcha_with_gemini(driver, wait):
            print("Failed to solve captcha")
            input("Press Enter to try again...")
            return

        # Click search with delay
        print("Clicking search button...")
        search_button = wait.until(
            EC.element_to_be_clickable((By.XPATH, "/html/body/div[1]/div[2]/div[1]/form/div[2]/div[2]/div[9]/div[2]/button[1]"))
        )
        time.sleep(2)
        search_button.click()
        
        # Add longer wait for search results to load completely
        print("Waiting for search results to load (60 seconds)...")
        time.sleep(60)  # Wait for 60 seconds after search button click
        
        print("Starting PDF downloads...")
        start_from = config["downloaded_pdf_number"] + 1
        end_at = config["pdf_range"]["end_serial"]

        for i in range(start_from, end_at + 1):
            try:
                print(f"Processing PDF {i}...")
                time.sleep(2)
                
                # Collect case details first
                case_data = {}
                try:
                    case_data['case_no'] = wait.until(
                        EC.presence_of_element_located((By.XPATH, f"/html/body/div[1]/div[2]/div[1]/div[4]/div/div/div/div[2]/div/div[2]/table/tbody/tr[{i}]/td[3]/button/u"))
                    ).text.strip()
                    
                    case_data['year'] = wait.until(
                        EC.presence_of_element_located((By.XPATH, f"/html/body/div[1]/div[2]/div[1]/div[4]/div/div/div/div[2]/div/div[2]/table/tbody/tr[{i}]/td[4]/button/u"))
                    ).text.strip()
                    
                    case_data['decision_date'] = wait.until(
                        EC.presence_of_element_located((By.XPATH, f"/html/body/div[1]/div[2]/div[1]/div[4]/div/div/div/div[2]/div/div[2]/table/tbody/tr[{i}]/td[9]"))
                    ).text.strip()
                    
                    case_data['judge_name'] = wait.until(
                        EC.presence_of_element_located((By.XPATH, f"/html/body/div[1]/div[2]/div[1]/div[4]/div/div/div/div[2]/div/div[2]/table/tbody/tr[{i}]/td[8]"))
                    ).text.strip()
                    
                    parties = wait.until(
                        EC.presence_of_element_located((By.XPATH, f"/html/body/div[1]/div[2]/div[1]/div[4]/div/div/div/div[2]/div/div[2]/table/tbody/tr[{i}]/td[6]"))
                    ).text.strip()
                    case_data['case_title'] = parties
                    
                except Exception as e:
                    print(f"Error collecting case details for row {i}: {e}")
                    continue
                
                # Get list of files before download
                download_dir = Path(config["download_directory"])
                files_before = set(download_dir.glob('*.pdf'))
                
                # Before clicking PDF button, remove any blocking elements
                remove_blocking_elements(driver)
                
                # Click PDF button
                pdf_xpath = f"/html/body/div[1]/div[2]/div[1]/div[4]/div/div/div/div[2]/div/div[2]/table/tbody/tr[{i}]/td[15]/button"
                pdf_button = wait.until(EC.element_to_be_clickable((By.XPATH, pdf_xpath)))
                
                # Scroll the button into view and ensure it's clickable
                driver.execute_script("arguments[0].scrollIntoView({block: 'center'});", pdf_button)
                time.sleep(1)
                
                # Try JavaScript click if regular click fails
                try:
                    pdf_button.click()
                except:
                    driver.execute_script("arguments[0].click();", pdf_button)
                
                time.sleep(2)
                
                # Check for record not found popup
                try:
                    ok_button = WebDriverWait(driver, 3).until(
                        EC.element_to_be_clickable((By.XPATH, "/html/body/div[9]/div/div[6]/button[1]"))
                    )
                    ok_button.click()
                    time.sleep(1)
                    print(f"Record not found for PDF {i}, clicked OK")
                    
                    # Update Excel with NOT AVAILABLE status
                    case_data['pdf_status'] = 'NOT AVAILABLE'
                    case_data['original_filename'] = ''
                    case_data['new_filename'] = ''
                    update_excel(config, i, case_data)
                    continue
                    
                except:
                    # No popup found, proceed with download and renaming
                    try:
                        # Wait for download to complete (increased max time to 45 seconds)
                        download_complete = False
                        max_wait = 45  # Increased from 30 to 45 seconds
                        print(f"Waiting for PDF {i} to download...")
                        
                        while max_wait > 0 and not download_complete:
                            time.sleep(2)  # Check every 2 seconds instead of 1
                            files_after = set(download_dir.glob('*.pdf'))
                            new_files = files_after - files_before
                            if new_files:
                                print(f"PDF {i} download detected")
                                download_complete = True
                                # Add extra wait to ensure file is completely written
                                time.sleep(3)  # Wait 3 more seconds after detection
                                break
                            max_wait -= 2
                            if max_wait % 10 == 0:  # Log progress every 10 seconds
                                print(f"Still waiting for PDF {i}... {max_wait} seconds remaining")
                        
                        if download_complete:
                            # Get the new file
                            new_file = list(new_files)[0]
                            original_filename = new_file.name
                            
                            # Construct new filename
                            parties_split = parties.split(" VS ", 1)
                            first_party = parties_split[0].strip() if len(parties_split) > 0 else "PARTY1"
                            second_party = parties_split[1].strip() if len(parties_split) > 1 else "PARTY2"
                            
                            new_filename = f"KAHC_{case_data['case_no']}_{case_data['year']}_{first_party}_VS_{second_party}.pdf"
                            new_filename = "".join(c for c in new_filename if c.isalnum() or c in "._- ")
                            
                            # Ensure the source file exists and is complete
                            old_path = download_dir / original_filename
                            new_path = download_dir / new_filename
                            
                            # Wait for file to be ready for renaming
                            time.sleep(2)
                            
                            try:
                                old_path.rename(new_path)
                                print(f"Successfully renamed {original_filename} to {new_filename}")
                                
                                # Update Excel with successful download
                                case_data['pdf_status'] = 'DOWNLOADED'
                                case_data['original_filename'] = original_filename
                                case_data['new_filename'] = new_filename
                                update_excel(config, i, case_data)
                                
                                # Update config
                                config["downloaded_pdf_number"] = i
                                save_config(config)
                            except Exception as rename_error:
                                print(f"Error renaming file: {rename_error}")
                                raise
                        else:
                            print(f"Timeout waiting for PDF download for row {i}")
                            raise Exception("Download timeout")
                            
                    except Exception as e:
                        print(f"Error processing download/rename for PDF {i}: {e}")
                        continue

                # Check if we've reached the end serial
                if i >= config["pdf_range"]["end_serial"]:
                    print(f"\nReached end serial number {config['pdf_range']['end_serial']}")
                    print("Saving final Excel updates...")
                    
                    # Get Excel workbook and save
                    try:
                        wb = load_workbook(config["excel_path"])
                        wb.save(config["excel_path"])
                        print(f"Excel file saved successfully: {config['excel_path']}")
                    except Exception as excel_error:
                        print(f"Error saving final Excel update: {excel_error}")
                    
                    print("\nScript completed successfully!")
                    return  # Exit the function after reaching end serial

            except Exception as e:
                print(f"Error downloading PDF {i}: {e}")
                continue

    except Exception as e:
        print(f"Error in main process: {e}")
        input("Press Enter to close the browser...")
    finally:
        if 'driver' in locals():
            driver.quit()

if __name__ == "__main__":
    print("Script starting...")
    main()
    print("Script finished.")
    input("Press Enter to exit...")
